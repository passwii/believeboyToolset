from flask import (
    Blueprint,
    request,
    send_file,
    jsonify,
    render_template,
    flash,
    redirect,
    url_for,
)
import io
import os
import csv
import openpyxl
from copy import copy
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import datetime
import pandas as pd
from core.log_service import LogService

yumai_analysis_bp = Blueprint("yumai_analysis", __name__)


@yumai_analysis_bp.route("/yumai-analysis", methods=["GET"])
def yumai_analysis_page():
    """渲染优麦云商品分析页面"""
    current_dir = os.path.dirname(__file__)
    projects_csv_path = os.path.join(current_dir, "model_file", "projects.csv")

    projects = []
    try:
        if os.path.exists(projects_csv_path):
            with open(projects_csv_path, "r", encoding="utf-8") as f:
                reader = csv.reader(f)
                header = next(reader, None)
                if header and header[0] == "项目名称":
                    for row in reader:
                        if row and row[0].strip():
                            projects.append(row[0])
                else:
                    if header and header[0].strip():
                        projects.append(header[0])
                    for row in reader:
                        if row and row[0].strip():
                            projects.append(row[0])
    except Exception as e:
        print(f"读取项目列表失败: {e}")
        projects = []

    return render_template(
        "data-analysis/product_analysis_yumai.html", projects=projects
    )


def allowed_file(filename):
    """检查文件是否为允许的格式"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"xlsx", "txt"}


def process_yumai_data(yumai_report_path, fba_report_path=None):
    """处理优麦云数据并可选择性地添加库存详情"""
    if not os.path.exists(yumai_report_path):
        raise FileNotFoundError("优麦云报表文件不存在！")

    wb = openpyxl.load_workbook(yumai_report_path)
    ws = wb.active

    try:
        header = {cell.value: cell.column for cell in ws[1]}
    except IndexError:
        raise ValueError("无法读取文件头部，请确保文件格式正确且至少有一行。")

    source_columns = [
        "SKU",
        "ASIN",
        "币种",
        "销量",
        "销售额",
        "可售",
        "广告花费",
        "广告曝光量",
        "广告点击量",
        "ACoAS",
    ]
    missing = [col for col in source_columns if col not in header]
    if missing:
        raise ValueError(f"上传文件缺少必需列: {', '.join(missing)}")

    ordered_columns = [
        "SKU",
        "ASIN",
        "币种",
        "销量",
        "销售额",
        "可售",
        "广告花费",
        "广告曝光量",
        "广告点击量",
        "广告占比",
    ]

    # 创建一个新的临时工作簿来处理主数据，以免损坏原始样式
    processed_wb = openpyxl.Workbook()
    new_ws = processed_wb.active
    new_ws.title = "优麦云分析"

    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    header_fill = PatternFill(fill_type="solid", fgColor="7A6AFF")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for i, col_name in enumerate(ordered_columns, 1):
        cell = new_ws.cell(row=1, column=i, value=col_name)
        cell.font = copy(header_font)
        cell.alignment = copy(header_alignment)
        cell.fill = copy(header_fill)
        cell.border = border

    source_col_indices = {name: header[name] for name in source_columns}
    new_col_indices = {name: i + 1 for i, name in enumerate(source_columns)}

    summary = {
        "销量": 0,
        "销售额": 0,
        "广告花费": 0,
        "广告曝光量": 0,
        "广告点击量": 0,
        "可售": 0,
    }
    numeric_cols = ["销量", "销售额", "广告花费", "广告曝光量", "广告点击量", "可售"]

    data_rows = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_name in source_columns:
            old_col_idx = source_col_indices[col_name]
            old_cell = ws.cell(row=row_idx, column=old_col_idx)
            row_data[col_name] = old_cell.value
        data_rows.append(row_data)

        for col_name in numeric_cols:
            val = row_data.get(col_name)
            if isinstance(val, (int, float)):
                summary[col_name] += val

    for r_idx, row_data in enumerate(data_rows, 2):
        for col_name in source_columns:
            new_col_idx = new_col_indices[col_name]
            new_cell = new_ws.cell(
                row=r_idx, column=new_col_idx, value=row_data[col_name]
            )
            new_cell.alignment = Alignment(horizontal="center", vertical="center")
            new_cell.border = border

    summary_row_idx = new_ws.max_row + 1
    summary_row_data = {
        "SKU": "汇总",
        "ASIN": "",
        "币种": "",
        "销量": summary["销量"],
        "销售额": summary["销售额"],
        "可售": summary["可售"],
        "广告花费": summary["广告花费"],
        "广告曝光量": summary["广告曝光量"],
        "广告点击量": summary["广告点击量"],
        "广告占比": "",
    }
    if summary["销售额"] and summary["销售额"] != 0:
        summary_row_data["广告占比"] = (
            f"{round(summary['广告花费'] / summary['销售额'] * 100, 2)}%"
        )

    summary_font = Font(bold=True)
    for i, col_name in enumerate(ordered_columns, 1):
        cell = new_ws.cell(
            row=summary_row_idx, column=i, value=summary_row_data.get(col_name)
        )
        cell.font = copy(summary_font)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # 如果有库存文件，添加库存详情sheet
    if fba_report_path and os.path.exists(fba_report_path):
        ws_inv = processed_wb.create_sheet("库存详情")
        inv_df = pd.read_csv(fba_report_path, sep="\t", encoding="utf-8")

        keep_cols = [
            "sku",
            "asin",
            "available",
            "inv-age-0-to-90-days",
            "inv-age-91-to-180-days",
            "inv-age-181-to-270-days",
            "inv-age-271-to-365-days",
            "inv-age-365-plus-days",
            "recommended-action",
        ]
        inv_df = inv_df[[col for col in keep_cols if col in inv_df.columns]].copy()

        rename_dict = {
            "sku": "SKU",
            "asin": "ASIN",
            "available": "可售库存",
            "inv-age-0-to-90-days": "0-90天",
            "inv-age-91-to-180-days": "91-180天",
            "inv-age-181-to-270-days": "181-270天",
            "inv-age-271-to-365-days": "271-365天",
            "inv-age-365-plus-days": "365+天",
            "recommended-action": "库存建议",
        }
        inv_df.rename(columns=rename_dict, inplace=True)

        sales_df = pd.DataFrame(data_rows)[["SKU", "销量"]].copy()
        sales_df.rename(columns={"销量": "本周销量"}, inplace=True)
        sales_df["本周销量"] = pd.to_numeric(
            sales_df["本周销量"], errors="coerce"
        ).fillna(0)

        inv_df = pd.merge(inv_df, sales_df, on="SKU", how="left")
        inv_df["本周销量"] = inv_df["本周销量"].fillna(0)

        new_column_order = [
            "SKU",
            "ASIN",
            "本周销量",
            "可售库存",
            "0-90天",
            "91-180天",
            "181-270天",
            "271-365天",
            "365+天",
            "库存建议",
        ]
        inv_df = inv_df[[col for col in new_column_order if col in inv_df.columns]]

        summary_data = {
            "SKU": "汇总行",
            "ASIN": "",
            "本周销量": inv_df["本周销量"].sum(),
            "可售库存": inv_df["可售库存"].sum(),
            "0-90天": inv_df["0-90天"].sum(),
            "91-180天": inv_df["91-180天"].sum(),
            "181-270天": inv_df["181-270天"].sum(),
            "271-365天": inv_df["271-365天"].sum(),
            "365+天": inv_df["365+天"].sum(),
            "库存建议": "",
        }
        summary_df = pd.DataFrame([summary_data])
        inv_df = pd.concat([inv_df, summary_df], ignore_index=True)

        for r_idx, row in enumerate(
            dataframe_to_rows(inv_df, index=False, header=True), 1
        ):
            for c_idx, value in enumerate(row, 1):
                cell = ws_inv.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
                if r_idx == 1:
                    cell.font = copy(header_font)
                    cell.fill = copy(header_fill)

        for col in ws_inv.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws_inv.column_dimensions[column].width = adjusted_width

    return processed_wb


@yumai_analysis_bp.route("/yumai-analysis/upload-file", methods=["POST"])
def upload_file():
    """处理优麦云分析相关的文件上传"""
    try:
        if "file" not in request.files:
            return jsonify({"success": False, "error": "没有文件被上传"}), 400

        file = request.files["file"]
        project_name = request.form.get("project_name")
        file_type = request.form.get("file_type")

        if not project_name:
            return jsonify({"success": False, "error": "项目名称不能为空"}), 400

        if not file_type or file_type not in ["yumai_report", "fba_report"]:
            return jsonify({"success": False, "error": "文件类型无效"}), 400

        if file and allowed_file(file.filename):
            upload_folder = os.path.join(
                os.getcwd(), "project", project_name, "uploaded_files"
            )
            os.makedirs(upload_folder, exist_ok=True)

            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            original_filename = file.filename
            file_ext = os.path.splitext(original_filename)[1].lower()

            new_filename = f"{project_name}_{file_type}_{timestamp}{file_ext}"

            file_path = os.path.join(upload_folder, new_filename)
            file.save(file_path)

            LogService.log(
                action="上传优麦分析文件",
                resource="优麦分析",
                details=f"项目: {project_name}, 文件类型: {file_type}, 文件名: {original_filename}",
                log_type="user",
                level="info",
            )

            return jsonify(
                {
                    "success": True,
                    "file_path": file_path,
                    "filename": new_filename,
                    "original_filename": original_filename,
                }
            )
        else:
            LogService.log(
                action="上传优麦分析文件失败",
                resource="优麦分析",
                details=f"项目: {project_name}, 文件类型: {file_type}, 错误: 文件类型不支持",
                log_type="user",
                level="warning",
            )
            return jsonify({"success": False, "error": "文件类型不支持"}), 400

    except Exception as e:
        LogService.log(
            action="上传优麦分析文件异常",
            resource="优麦分析",
            details=f"项目: {request.form.get('project_name')}, 错误: {str(e)}",
            log_type="system",
            level="error",
        )
        return jsonify({"success": False, "error": str(e)}), 500


@yumai_analysis_bp.route("/yumai-analysis/submit", methods=["POST"])
def submit_yumai_analysis():
    """处理提交的表单并生成分析报告"""
    try:
        project_name = request.form.get("project_name")
        report_start_date = request.form.get("report_start_date")
        report_end_date = request.form.get("report_end_date")
        yumai_report_path = request.form.get("yumai_report_path")
        fba_report_path = request.form.get("fba_report_path")

        if not all(
            [project_name, report_start_date, report_end_date, yumai_report_path]
        ):
            flash("请确保所有必填项都已填写并上传了优麦云报表。")
            return redirect(url_for("yumai_analysis.yumai_analysis_page"))

        if not os.path.exists(yumai_report_path):
            flash("优麦云报表文件不存在，请重新上传。")
            return redirect(url_for("yumai_analysis.yumai_analysis_page"))

        if fba_report_path and not os.path.exists(fba_report_path):
            flash("库存报告文件不存在，将不包含库存分析。")
            fba_report_path = None

        result_wb = process_yumai_data(yumai_report_path, fba_report_path)

        output = io.BytesIO()
        result_wb.save(output)
        output.seek(0)

        start_date_part = report_start_date.replace("-", "")
        end_date_part = report_end_date.replace("-", "")[4:]
        report_date = f"{start_date_part}-{end_date_part}"
        filename = f"{project_name}_YumaiAnalysis_{report_date}.xlsx"

        LogService.log(
            action="生成优麦分析报告",
            resource="优麦分析",
            details=f"项目: {project_name}, 日期: {report_date}",
            log_type="user",
            level="info",
        )

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except ValueError as e:
        LogService.log(
            action="生成优-麦分析报告失败",
            resource="优麦分析",
            details=f"错误: {str(e)}",
            log_type="system",
            level="error",
        )
        flash(f"处理失败: {str(e)}")
        return redirect(url_for("yumai_analysis.yumai_analysis_page"))
    except Exception as e:
        LogService.log(
            action="生成优麦分析报告异常",
            resource="优麦分析",
            details=f"错误: {str(e)}",
            log_type="system",
            level="error",
        )
        flash(f"处理过程中发生意外错误: {str(e)}")
        return redirect(url_for("yumai_analysis.yumai_analysis_page"))
