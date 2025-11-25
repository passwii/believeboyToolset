from flask import Blueprint, render_template, request, send_file, redirect, flash, url_for
import os
import io
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from core.log_service import LogService

daily_report_bp = Blueprint('daily_report', __name__)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'txt', 'xlsx'}

def save_file(file, folder, filename):
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath

def process_daily_report(project_name, report_date, sales_report, ad_report, fba_report):
    current_time = datetime.datetime.now().strftime('%H-%M-%S')
    source_folder = os.getcwd()
    os.chdir(source_folder)

    project_folder_path = os.path.join(source_folder, 'project', project_name, '日报')
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, 'project', project_name, 'tmp')
    os.makedirs(tmp_folder_path, exist_ok=True)

    sales_report_path = f'{project_name}_sales_report_{report_date}.txt'
    ad_report_path = f'{project_name}_ad_report_{report_date}.xlsx'
    fba_report_path = f'{project_name}_fba_report_{report_date}.txt'

    with open(sales_report_path, 'wb') as f:
        f.write(sales_report.read())

    with open(ad_report_path, 'wb') as f:
        f.write(ad_report.read())

    with open(fba_report_path, 'wb') as f:
        f.write(fba_report.read())

    daily_sales = pd.read_csv(sales_report_path, sep='\t', encoding='utf-8')
    daily_ad_report = pd.read_excel(ad_report_path, engine='openpyxl')
    fba = pd.read_csv(fba_report_path, sep='\t', encoding='utf-8')

    files_to_copy = [
        (sales_report_path, f'{tmp_folder_path}/{project_name}_sales_report_{report_date}_{current_time}.txt'),
        (ad_report_path, f'{tmp_folder_path}/{project_name}_ad_report_{report_date}_{current_time}.xlsx'),
        (fba_report_path, f'{tmp_folder_path}/{project_name}_fba_report_{report_date}_{current_time}.txt')
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)
        os.remove(src)

    daily_sales = daily_sales.loc[daily_sales['order-status'].isin(['Pending', 'Shipped', 'Unshipped'])]
    daily_sales = daily_sales.groupby(['sku'])[['quantity', 'item-price']].sum().reset_index()
    daily_sales = daily_sales.rename(columns={'sku': 'SKU', 'quantity': '订单量', 'item-price': '销售额'})

    ad_columns_keep = ['广告SKU', '广告ASIN', '展示量', '点击量', '花费', '7天总销售量(#)']
    daily_ad_report = daily_ad_report[ad_columns_keep]
    daily_ad_report = daily_ad_report.rename(columns={'广告SKU': 'SKU', '广告ASIN': 'ASIN', '花费': '广告花费',
                                                      '展示量': '曝光量', '7天总销售量(#)': '广告订单'})
    daily_ad_report = (daily_ad_report.groupby(['SKU'])
                       .agg({'曝光量': 'sum', '点击量': 'sum', '广告花费': 'sum', '广告订单': 'sum', 'ASIN': 'first'})
                       .reset_index())
    daily_ad_report['单次点击花费'] = daily_ad_report['广告花费'] / daily_ad_report['点击量'] \
        if daily_ad_report['点击量'].sum() > 0 else 0

    fba_columns_keep = ['sku', 'available']
    fba = fba[fba_columns_keep]
    fba = fba.rename(columns={'sku': 'SKU', 'available': '可售库存'})

    merged_data = pd.merge(daily_sales, daily_ad_report, on='SKU', how='outer')
    merged_data = pd.merge(merged_data, fba, on='SKU', how='outer')
    merged_data = merged_data.fillna(0)
    # 按SKU列排序
    merged_data = merged_data.sort_values('SKU')

    overview_data = {
        '日期(US)': [],
        'SKU': [],
        'ASIN': [],
        '订单量': [],
        '销售额': [],
        '曝光量': [],
        '点击量': [],
        '单次点击花费': [],
        '广告花费': [],
        '广告订单': [],
        '可售库存': []
    }
    df_overview = pd.DataFrame(overview_data)
    df_overview = pd.concat([df_overview, merged_data], ignore_index=True)
    df_overview = df_overview.fillna(0)
    df_overview['日期(US)'] = report_date

    for column in ['订单量', '曝光量', '点击量', '广告订单', '可售库存']:
        df_overview[column] = df_overview[column].astype(int)
    for column in ['单次点击花费', '广告花费']:
        df_overview[column] = df_overview[column].round(2)
    # 模板文件路径
    template_file = os.path.join('apps', 'model_file', 'daily_template.xlsx')
    # 日报文件路径
    project_daily_file_path = os.path.join(project_folder_path, f'{project_name}_{report_date}_日报.xlsx')
    # 加载模板文件
    workbook = load_workbook(template_file)
    ws = workbook.active

    for r_idx, row in enumerate(dataframe_to_rows(df_overview, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(project_daily_file_path)

    with open(project_daily_file_path, 'rb') as f:
        file_content = f.read()

    return file_content, f'{project_name}_{report_date}_日报.xlsx'

@daily_report_bp.route('/daily-report', methods=['GET', 'POST'])
def daily_report():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        report_date = request.form.get('report_date')
        sales_report = request.files.get('sales_report')
        ad_report = request.files.get('ad_report')
        fba_report = request.files.get('fba_report')

        if not (project_name and report_date and sales_report and ad_report and fba_report):
            flash('请填写所有字段并上传所有文件')
            return redirect(url_for('dataset.daily_report_page'))

        if not (allowed_file(sales_report.filename) and allowed_file(ad_report.filename) and allowed_file(
                fba_report.filename)):
            flash('文件格式不正确')
            return redirect(url_for('dataset.daily_report_page'))

        try:
            file_content, filename = process_daily_report(project_name, report_date, sales_report, ad_report, fba_report)
            
            # 记录生成日报成功日志
            LogService.log(
                action="生成日报",
                resource="日报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 文件: {filename}",
                log_type="user",
                level="info"
            )
            
            # 手动创建响应以避免Flask send_file的中文文件名问题
            from flask import make_response
            import urllib.parse
            
            response = make_response(file_content)
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            # 只使用filename*参数，不使用filename参数，避免浏览器选择截断的文件名
            encoded_filename = urllib.parse.quote(filename)
            response.headers['Content-Disposition'] = f"attachment; filename*=UTF-8''{encoded_filename}"
            
            return response
        except Exception as e:
            # 记录生成日报失败日志
            LogService.log(
                action="生成日报失败",
                resource="日报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 错误: {str(e)}",
                log_type="user",
                level="error"
            )
            flash(f'生成日报时发生错误: {str(e)}', 'error')
            return redirect(url_for('dataset.daily_report_page'))
    
    return render_template('data-analysis/daily_report.html')
