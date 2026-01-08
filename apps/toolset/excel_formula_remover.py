import openpyxl
import io

class ExcelFormulaRemover:
    """
    A class to remove formulas from an .xlsx file while preserving values and styles.
    """
    def __init__(self, file_stream):
        """
        Initializes the remover with a file stream.
        
        Args:
            file_stream (io.BytesIO): A file-like object containing the xlsx data.
        """
        if not hasattr(file_stream, 'read'):
            raise TypeError("Input must be a file-like object (e.g., io.BytesIO)")
        self.file_stream = file_stream

    def process(self):
        """
        Removes all formulas from the Excel workbook, replacing them with their last calculated values.
        
        Returns:
            io.BytesIO: A new BytesIO stream containing the processed workbook.
        
        Raises:
            Exception: If there is an error during file processing.
        """
        try:
            # The input stream can only be read once. We need to read its content
            # into memory to be able to load it multiple times.
            file_content = self.file_stream.read()

            # --- First pass: Load with formulas to identify which cells to change ---
            # We need a new BytesIO stream for each load.
            workbook_with_formulas = openpyxl.load_workbook(io.BytesIO(file_content))
            
            # --- Second pass: Load in data_only mode to get calculated values ---
            workbook_with_values = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)

            # Iterate through both workbooks simultaneously and copy values where formulas exist
            for sheet_name in workbook_with_formulas.sheetnames:
                sheet_with_formulas = workbook_with_formulas[sheet_name]
                sheet_with_values = workbook_with_values[sheet_name]

                # Iterate through all cells in the sheet that contains formulas
                for row_idx, row in enumerate(sheet_with_formulas.iter_rows(), 1):
                    for col_idx, cell in enumerate(row, 1):
                        # If the cell in the original workbook has a formula...
                        if cell.data_type == 'f':
                            # ...get the value from the same cell in the data_only workbook...
                            value_cell = sheet_with_values.cell(row=row_idx, column=col_idx)
                            # ...and assign its value to the cell in the formula workbook.
                            cell.value = value_cell.value
            
            # Save the modified workbook (the one that originally had formulas)
            output_stream = io.BytesIO()
            workbook_with_formulas.save(output_stream)
            output_stream.seek(0)
            
            return output_stream
        except Exception as e:
            # It's good practice to log the error
            print(f"Error processing Excel file: {e}")
            # Re-raise the exception to be handled by the caller
            raise

def process_excel_file_stream(input_stream):
    """
    A convenience function to instantiate the class and process the file.
    
    Args:
        input_stream (io.BytesIO): The input Excel file stream.
        
    Returns:
        io.BytesIO: The processed Excel file stream.
    """
    remover = ExcelFormulaRemover(file_stream=input_stream)
    return remover.process()
