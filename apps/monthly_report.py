from flask import Blueprint, render_template, request, send_file, redirect, flash, url_for
import os
import io
import datetime
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from core.log_service import LogService

monthly_report_bp = Blueprint('monthly_report', __name__)

def reset_style(filename):
    wb = load_workbook(filename)
    line_t = Side(style='thin', color='000000')

    sty1 = NamedStyle(name='sty1')
    sty1.font = Font(name='等线', size=12, b=True)
    sty1.alignment = Alignment(horizontal='left', vertical='center')
    sty1.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    sty2 = NamedStyle(name='sty2')
    sty2.font = Font(name='等线', size=12)
    sty2.alignment = Alignment(horizontal='left', vertical='center')
    sty2.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    sty3 = NamedStyle(name='sty3')
    sty3.font = Font(name='等线', size=12)
    sty3.alignment = Alignment(horizontal='center', vertical='center')
    sty3.border = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    for ws1 in wb:
        nrows = ws1.max_row
        ncols = ws1.max_column
        width = 12
        height = 15
        for r in range(1, nrows + 1):
            for c in range(1, ncols + 1):
                if r == 1:
                    ws1.cell(r, c).style = sty1
                    ws1.row_dimensions[r].height = height
                else:
                    ws1.cell(r, c).style = sty2
                    ws1.column_dimensions[get_column_letter(
                        c)].width = width
        #冻结首行
        ws1.freeze_panes = ws1['A2']
    
    # 
    wb['报表总览'].sheet_properties.tabColor = "1072BA"
    wb['交易一览'].sheet_properties.tabColor = "CCCCCC" # 深灰色
    
    
    return wb.save(filename)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv'}

def save_file(file, folder, filename):
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath

def process_monthly_report(project_name, report_date, payment_range_report):
    current_time = datetime.datetime.now().strftime('%H-%M-%S')
    source_folder = os.getcwd()
    os.chdir(source_folder)

    report_name = f'{project_name}_美国站_{report_date}'

    project_folder_path = os.path.join(source_folder, 'project', project_name, '月报')
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, 'project', project_name, 'tmp')
    os.makedirs(tmp_folder_path, exist_ok=True)

    payment_range_report_path = f'{project_name}_payment_range_report_{report_date}.csv'

    with open(payment_range_report_path, 'wb') as f:
        f.write(payment_range_report.read())

    payment_range_report = pd.read_csv(payment_range_report_path, thousands=',', skiprows=7, encoding='utf-8')

    files_to_copy = [
        (payment_range_report_path, f'{tmp_folder_path}/{project_name}_payment_range_report_{report_date}_{current_time}.csv')
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)
        os.remove(src)
    PRR = payment_range_report
    PRR.fillna(0, inplace=True)
    PRR['quantity'] = PRR['quantity'].astype(int)
    
    # DataFrame - Type 一级分类
    Order = PRR.loc[PRR['type'].isin(['Order'])]
    Refund = PRR.loc[PRR['type'].isin(['Refund'])]
    #空白 Type 分类
    Blank = PRR.loc[PRR['type'] == 0]
    
    Chargeback_Refund = PRR.loc[PRR['type'].isin(['Chargeback Refund'])]
    Liquidation = PRR.loc[PRR['type'].isin(['Liquidations'])]
    Adjustment = PRR.loc[PRR['type'].isin(['Adjustment'])]
    Fee_Adjustment = PRR.loc[PRR['type'].isin(['Fee Adjustment'])]
    Service_Fee = PRR.loc[PRR['type'].isin(['Service Fee'])]
    FBA_Inventory_Reimbursement = PRR.loc[PRR['type'].isin(['FBA Inventory Reimbursement'])]
    FBA_Inventory_Fee = PRR.loc[PRR['type'].isin(['FBA Inventory Fee'])]
    Amazon_Fees = PRR.loc[PRR['type'].isin(['Amazon Fees'])]
    # Fee Adjustment 亚马逊 返还 的派送费(产品发货尺寸, 重新测量的费用) 
    Fee_Adjustment = PRR.loc[PRR['type'].isin(['Fee Adjustment'])]
    LD = PRR.loc[PRR['type'].isin(['Deal Fee'])]
    Debt = PRR.loc[PRR['type'].isin(['Debt'])]
    FBA_Transaction_fees = PRR.loc[PRR['type'].isin(['FBA Transaction fees'])]

    
    # DataFrame - fulfillment / Description 二级分类 
    FBA_Order = Order.loc[Order['fulfillment'].isin(['Amazon'])]
    FBM_Order = Order.loc[Order['fulfillment'].isin(['Seller'])]
    
    FBA_Refund = Refund.loc[Refund['fulfillment'].isin(['Amazon'])]
    FBM_Refund = Refund.loc[Refund['fulfillment'].isin(['Seller'])]
    
    #Income赔偿
    # 不是FBA Inventory Reimbursement - General Adjustment 的其他赔偿
    Adjustment_Income = Adjustment.loc[~Adjustment['description'].isin(['FBA Inventory Reimbursement - General Adjustment'])]
    #Expense赔偿
    Adjustment_Expense = Adjustment.loc[Adjustment['description'].isin(['FBA Inventory Reimbursement - General Adjustment'])]
    
    Advertising = Service_Fee.loc[Service_Fee['description'].isin(['Cost of Advertising'])]
    Subscription = Service_Fee.loc[Service_Fee['description'].isin(['Subscription'])]
    FBA_Inbound = Service_Fee.loc[Service_Fee['description'].isin(['FBA Inbound Placement Service Fee'])]
    AGL_Selection = Service_Fee.loc[Service_Fee['description'].isin([
        'FBA International Freight', 
        'FBA International Freight Duties and Taxes Charge'
        ])]
    # AmazonFees 优惠券 + 一般服务费（不含广告）
    Amazon_Fees_and_Service_Fee_without_AD = PRR.loc[(PRR['type'].isin(['Amazon Fees', 'Service Fee'])) & (PRR['description'] != 'Cost of Advertising') & (PRR['description'] != 'Refund for Advertiser')]
    
    
    Service_Refund_for_Advertiser = Service_Fee.loc[Service_Fee['description'].isin(['Refund for Advertiser'])]
    # 空白, Description 二级分类包含Price Discount开头的项
    Price_Discount = Blank.loc[Blank['description'].str.startswith('Price Discount')]
    
    # 重新测量费的计算口径需要未来确认
    Product_sales_non_FBA = round(FBM_Order['product sales'].sum(), 2)
    Product_sale_refunds_non_FBA = round(FBM_Refund['product sales'].sum(), 2)
    FBA_product_sales = round(FBA_Order['product sales'].sum(), 2)
    FBA_product_sale_refunds = round(FBA_Refund['product sales'].sum(), 2) + round(FBA_Refund['other'].sum(), 2) # FBA 退款的其他费用也计入退款
    Fee_Adjustment_fee = round(Fee_Adjustment.loc[Fee_Adjustment['description'].isin(['FBA Weight/Dimension Change'])]['total'].sum(), 2)
    FBA_inventory_credit = round(Adjustment_Income['total'].sum(), 2) + Fee_Adjustment_fee
    FBA_liquidation_proceeds = round(Liquidation['product sales'].sum(), 2)
    Shipping_credits = round(Order['shipping credits'].sum(), 2)
    Shipping_credit_refunds = round(Refund['shipping credits'].sum(), 2)
    Gift_wrap_credits = round(Order['gift wrap credits'].sum(), 2)
    Gift_wrap_credits_refunds = round(Refund['gift wrap credits'].sum(), 2)
    Promotional_rebates = round(Order['promotional rebates'].sum(), 2)
    Promotional_rebate_refunds = round(Refund['promotional rebates'].sum(), 2)
    A_to_z_Guarantee_claims = 0
    Chargebacks = round(Chargeback_Refund['total'].sum(), 2)
    SAFE_T_reimbursement = 0

    Income_total = round(Product_sales_non_FBA + Product_sale_refunds_non_FBA + FBA_product_sales + FBA_product_sale_refunds + FBA_inventory_credit + FBA_liquidation_proceeds + Shipping_credits + Shipping_credit_refunds + Gift_wrap_credits + Gift_wrap_credits_refunds + Promotional_rebates + Promotional_rebate_refunds + Chargebacks, 2)
    
    '''计算 Expense (与PDF完全一致) '''
    Seller_fulfilled_selling_fees = round(FBM_Order['selling fees'].sum(), 2)
    FBA_selling_fees = round(FBA_Order['selling fees'].sum(), 2)
    Selling_fee_refunds = round(Refund['selling fees'].sum(), 2) #目前不确定是否是全部还是Order和Refund分别计算
    FBA_transaction_fees = round(FBA_Order['fba fees'].sum(), 2) + round(FBA_Transaction_fees['fba fees'].sum(), 2)
    FBA_transaction_fee_refunds = round(FBA_Refund['fba fees'].sum(), 2)
    Other_transaction_fees = 0
    Other_transaction_fee_refunds = 0
    FBA_inventory_and_inbound_services_fees = round(FBA_Inventory_Fee['total'].sum(), 2)
    Shipping_label_purchases = 0
    Shipping_label_refunds = 0
    Carrier_shipping_label_adjustments = 0
    # 服务费 = 优惠券 + 一般服务费（不含广告）+ 闪电促销费 + Price Discount
    Coupon_fee = round(Amazon_Fees['total'].sum(), 2)
    Lightning_Deal_Fee = round(LD['total'].sum(), 2)
    Service_Price_Discount = round(Price_Discount['total'].sum(), 2)
    # 服务费去掉广告费用的综合 + 优惠券就是整体的PDF服务费计算
    Service_fees_without_ADs = round(Service_Fee.loc[~Service_Fee['description'].isin(['Cost of Advertising', 'Refund for Advertiser'])]['total'].sum(), 2)
    Service_fees = Coupon_fee + Service_fees_without_ADs + Lightning_Deal_Fee + Service_Price_Discount
    
    
    # 退款管理费和佣金退款合并计算在PDF中是和佣金退款合并的，所以是0
    Refund_administration_fees = 0
    Adjustments = round(Adjustment_Expense['total'].sum(), 2)
    Cost_of_Advertising = round(Advertising['total'].sum(), 2)
    Refund_for_Advertiser = round(Service_Refund_for_Advertiser['total'].sum(), 2)
    # Liquidation 清算费用
    Liquidations_fees = round(Liquidation['other transaction fees'].sum(), 2)
    Receivables = 0
    Deductions = 0
    Amazon_Shipping_Charge_Adjustments = 0

    Expense_total = round(Seller_fulfilled_selling_fees + FBA_selling_fees + Selling_fee_refunds + FBA_transaction_fees + FBA_transaction_fee_refunds + Other_transaction_fees + Other_transaction_fee_refunds + FBA_inventory_and_inbound_services_fees + Shipping_label_purchases + Shipping_label_refunds + Carrier_shipping_label_adjustments + Service_fees + Refund_administration_fees + Adjustments + Cost_of_Advertising + Refund_for_Advertiser + Liquidations_fees + Receivables + Deductions + Amazon_Shipping_Charge_Adjustments, 2)

    # 销售SKU明细
    skuGroup = Order.groupby(['sku'], as_index=False).agg({'quantity': 'sum'})
    Order_QTY = int(Order['quantity'].sum())

    # 退款SKU明细
    refund_skuGroup = Refund.groupby(['sku'], as_index=False).agg({'quantity': 'sum'})
    Refund_QTY = int(Refund['quantity'].sum())

    # 平台推广费用
    spFee = Cost_of_Advertising
    spCount = format(abs(spFee) / Income_total, '.2%')

    # 信用卡扣费
    AMZ_Card = round(Debt['total'].sum(), 2)

    # 处理子项求和筛选展示
    
    
    
    # 汇总数据
    pt1 = skuGroup.sort_values(by='quantity', ascending=False, inplace=False)
    pt2 = refund_skuGroup.sort_values(by='quantity', ascending=False, inplace=False)
    
    pt3 = pd.DataFrame({
    'Income':[
        'Product sales (non-FBA)',
        'Product sale refunds (non-FBA) ',
        'FBA product sales',
        'FBA product sale refunds',
        'FBA inventory credit',
        'FBA liquidation proceeds',
        'Shipping credits',
        'Shipping credit refunds',
        'Gift wrap credits',
        'Gift wrap credit refunds',
        'Promotional rebates',
        'Promotional rebate refunds',
        'A-to-z Guarantee claims',
        'Chargebacks',
        'SAFE-T reimbursement',
        '',
        '',
        '',
        '',
        '',
        '',
        'Total_Income'],
    
    '收入':[
        '销售额（非FBA）',
        '销售额退款（非FBA）', 
        '销售额（FBA）', 
        '销售额退款（FBA）', 
        'FBA库存赔偿（FBA）', 
        'FBA清算收入', 
        '运费收入', 
        '运费退款', 
        '礼品包装收入', 
        '礼品包装退款', 
        '促销折扣', 
        '促销折扣退款', 
        'A-to-z保障索赔',
        '拒付退款', 
        'SAFE-T赔偿',
        '',
        '',
        '',
        '',
        '',
        '',
        '合计销售额'],
    
    'In金额（USD）':[
        Product_sales_non_FBA, 
        Product_sale_refunds_non_FBA, 
        FBA_product_sales, 
        FBA_product_sale_refunds, 
        FBA_inventory_credit, 
        FBA_liquidation_proceeds, 
        Shipping_credits, 
        Shipping_credit_refunds, 
        Gift_wrap_credits, 
        Gift_wrap_credits_refunds, 
        Promotional_rebates, 
        Promotional_rebate_refunds, 
        A_to_z_Guarantee_claims,
        Chargebacks, 
        SAFE_T_reimbursement,
        '',
        '',
        '',
        '',
        '',
        '',
        Income_total
        ],
    
    'In源表':[
        'FBM 订单-订单金额', 
        'FBM 退款-订单金额', 
        'FBA 订单-订单金额', 
        'FBA 退款-订单金额', 
        'FBA库存赔偿-订单金额+其他金额', 
        '清算费用-订单金额', 
        '所有订单-运费金额', 
        '所有退款-运费金额', 
        '所有订单-礼品包装金额', 
        '所有退款-礼品包装金额', 
        '所有订单-促销折扣金额', 
        '所有退款-促销折扣金额', 
        '', 
        '拒付退款-总计',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        ''
        ],
    
    '':[
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        ''
        ],

    'Expense':[
        'Seller fulfilled selling fees', 
        'FBA selling fees', 
        'Selling fee refunds', 
        'FBA transaction fees', 
        'FBA transaction fee refunds', 
        'Other transaction fees', 
        'Other transaction fee refunds', 
        'FBA inventory and inbound services fees', 
        'Shipping label purchases', 
        'Shipping label refunds', 
        'Carrier shipping label adjustments', 
        'Service fees', 
        'Refund administration fees', 
        'Adjustments',
        'Cost of Advertising', 
        'Refund for Advertiser',
        'Liquidations fees',
        'Receivables',
        'Deductions',
        'Amazon Shipping Charge Adjustments',
        '',
        'Total_Expense'
        ],
    
    '支出':[
        '卖家自配送销售佣金', 
        'FBA销售佣金',
        '销售佣金退款',
        'FBA派送费', 
        'FBA派送费退款',
        '其他交易费用',
        '其他交易费用退款',
        'FBA库存和入库服务费',
        '购买配送标签',
        '配送标签退款',
        '承运商配送标签调整',
        '服务费',
        '退款管理费',
        '调整项',
        '广告成本',
        '广告商退款',
        '清算处理费',
        '应收账款',
        '扣款',
        '亚马逊配送运费调整',
        '',
        '合计费用'
        ],
    
    'Ex金额（USD）':[
        Seller_fulfilled_selling_fees,
        FBA_selling_fees,
        Selling_fee_refunds,
        FBA_transaction_fees,
        FBA_transaction_fee_refunds,
        Other_transaction_fees,
        Other_transaction_fee_refunds,
        FBA_inventory_and_inbound_services_fees,
        Shipping_label_purchases,
        Shipping_label_refunds,
        Carrier_shipping_label_adjustments,
        Service_fees,
        Refund_administration_fees,
        Adjustments,
        Cost_of_Advertising,
        Refund_for_Advertiser,
        Liquidations_fees,
        Receivables,
        Deductions,
        Amazon_Shipping_Charge_Adjustments,
        '',
        Expense_total
        ],
    
    'Ex源表':[
        'FBM 订单-销售佣金',
        'FBA 订单-销售佣金', 
        '所有退款-销售佣金 + 退款管理费',
        'FBA 订单 - 派送费',
        'FBA 退款 - 派送费',
        '', 
        '', 
        'FBA仓储及入库服务费 - 总计',
        '', 
        '',
        '',
        '服务费（不含广告）-总计',
        '合并在销售佣金退款中扣除', 
        '其他赔偿调整-总计',
        '广告费-总计',        
        '广告退款-总计', 
        '清算费用-其他交易费用',
        '',
        '',
        '',
        '',
        ''
        ], 
    })  
    

    
    project_monthly_file_path = os.path.join(project_folder_path, f'{project_name}_{report_date}_monthly_{current_time}.xlsx')
    
    # 使用 ExcelWriter 的上下文管理器
    with pd.ExcelWriter(project_monthly_file_path, engine='xlsxwriter') as writer:
        # 创建一个字典来存储所有需要写入的数据帧和对应的sheet名称
        sheets_to_write = {
            '报表总览': pt3,
            '销售SKU明细':pt1,
            '退款SKU明细':pt2,
            '交易一览': PRR,
            '所有订单': Order,
            '所有退款': Refund,
            'FBM 订单':FBM_Order,
            'FBM 退款':FBM_Refund,
            'FBA 订单':FBA_Order,
            'FBA 退款':FBA_Refund,
            'FBA库存赔偿':Adjustment_Income,
            '其他赔偿':Adjustment_Expense,
            '清算费用':Liquidation, # Liquidation['product sales']
            '拒付退款':Chargeback_Refund,
            'FBA仓储及入库服务费':FBA_Inventory_Fee,
            '服务费（不含广告）':Amazon_Fees_and_Service_Fee_without_AD,
            '广告费':Advertising,
            '广告退款':Service_Refund_for_Advertiser,
        }
        
        # 使用循环一次性写入所有sheet
        for sheet_name, df in sheets_to_write.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # 重置样式
    reset_style(project_monthly_file_path)

    # 使用with语句读取文件内容
    with open(project_monthly_file_path, 'rb') as f:
        file_content = f.read()

    return file_content, f'月度财务报表_{report_name}.xlsx'

@monthly_report_bp.route('/monthly-report', methods=['GET', 'POST'])
def monthly_report():
    if request.method == 'POST':
        project_name = request.form.get('project_name')
        report_date = request.form.get('report_date')
        payment_range_report = request.files.get('payment_range_report')

        if not project_name or not report_date or not payment_range_report:
            flash('请填写所有必填项')
            return redirect(url_for('dataset.monthly_report'))

        if not allowed_file(payment_range_report.filename):
            flash('文件格式不正确')
            return redirect(url_for('dataset.monthly_report'))

        try:
            file_content, filename = process_monthly_report(project_name, report_date, payment_range_report)
            
            # 记录生成月报成功日志
            LogService.log(
                action="生成月报",
                resource="月报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 文件: {filename}",
                log_type="user",
                level="info"
            )
            
            # 创建响应对象
            response = send_file(
                io.BytesIO(file_content),
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # 添加响应头，通知前端重置表单
            response.headers['X-Form-Reset'] = 'true'
                
            return response
        except Exception as e:
            # 记录生成月报失败日志
            LogService.log(
                action="生成月报失败",
                resource="月报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 错误: {str(e)}",
                log_type="user",
                level="error"
            )
            flash(f'生成月报时发生错误: {str(e)}', 'error')
            return redirect(url_for('dataset.monthly_report'))

    return render_template('data-analysis/monthly_report.html')

