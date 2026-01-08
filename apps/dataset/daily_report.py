from flask import (
    Blueprint,
    render_template,
    request,
    send_file,
    redirect,
    flash,
    url_for,
)
import os
import io
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil
from core.log_service import LogService

daily_report_bp = Blueprint("daily_report", __name__)


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in {"txt", "xlsx"}


@daily_report_bp.route("/upload-file", methods=["POST"])
def upload_file():
    """处理单个文件上传的API端点"""
    try:
        if "file" not in request.files:
            return {"success": False, "error": "没有文件被上传"}, 400

        file = request.files["file"]
        project_name = request.form.get("project_name")
        file_type = request.form.get("file_type")

        if not project_name:
            return {"success": False, "error": "项目名称不能为空"}, 400

        if not file_type or file_type not in [
            "sales_report",
            "fba_report",
            "ad_report",
        ]:
            return {"success": False, "error": "文件类型无效"}, 400

        if file and allowed_file(file.filename):
            # 创建项目上传文件夹
            upload_folder = os.path.join(
                os.getcwd(), "project", project_name, "uploaded_files"
            )
            os.makedirs(upload_folder, exist_ok=True)

            # 生成文件名
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            original_filename = file.filename
            file_ext = os.path.splitext(original_filename)[1].lower()

            # 根据文件类型命名
            if file_type == "sales_report":
                new_filename = f"{project_name}_Sales_Report_{timestamp}{file_ext}"
            elif file_type == "fba_report":
                new_filename = f"{project_name}_FBA_Report_{timestamp}{file_ext}"
            elif file_type == "ad_report":
                new_filename = f"{project_name}_AD_Report_{timestamp}{file_ext}"

            file_path = os.path.join(upload_folder, new_filename)
            file.save(file_path)

            # 记录文件上传成功日志
            LogService.log(
                action="上传日报文件",
                resource="日报功能",
                details=f"项目: {project_name}, 文件类型: {file_type}, 文件名: {original_filename}",
                log_type="user",
                level="info",
            )

            print(f"文件已上传: {original_filename} -> {file_path}")

            return {
                "success": True,
                "file_path": file_path,
                "filename": new_filename,
                "original_filename": original_filename,
            }
        else:
            # 记录文件上传失败日志
            LogService.log(
                action="上传日报文件失败",
                resource="日报功能",
                details=f"项目: {project_name}, 文件类型: {file_type}, 错误: 文件类型不支持",
                log_type="user",
                level="warning",
            )
            return {"success": False, "error": "文件类型不支持"}, 400

    except Exception as e:
        # 记录文件上传异常日志
        LogService.log(
            action="上传日报文件异常",
            resource="日报功能",
            details=f"项目: {project_name}, 错误: {str(e)}",
            log_type="user",
            level="error",
        )
        print(f"文件上传错误: {str(e)}")
        import traceback

        traceback.print_exc()
        return {"success": False, "error": str(e)}, 500


def save_file(file, folder, filename):
    if not os.path.exists(folder):
        os.makedirs(folder)
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath


def process_daily_report_from_paths(
    project_name, report_date, sales_report_path, ad_report_path, fba_report_path
):
    """从文件路径处理日报"""
    current_time = datetime.datetime.now().strftime("%H-%M-%S")
    source_folder = os.getcwd()
    os.chdir(source_folder)

    project_folder_path = os.path.join(source_folder, "project", project_name, "日报")
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, "project", project_name, "tmp")
    os.makedirs(tmp_folder_path, exist_ok=True)

    # 复制已上传的文件到临时文件夹
    files_to_copy = [
        (
            sales_report_path,
            f"{tmp_folder_path}/{project_name}_sales_report_{report_date}_{current_time}.txt",
        ),
        (
            ad_report_path,
            f"{tmp_folder_path}/{project_name}_ad_report_{report_date}_{current_time}.xlsx",
        ),
        (
            fba_report_path,
            f"{tmp_folder_path}/{project_name}_fba_report_{report_date}_{current_time}.txt",
        ),
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)

    # 读取文件进行数据处理
    daily_sales = pd.read_csv(sales_report_path, sep="\t", encoding="utf-8")
    daily_ad_report = pd.read_excel(ad_report_path, engine="openpyxl")
    fba = pd.read_csv(fba_report_path, sep="\t", encoding="utf-8")

    # 数据处理逻辑（与原函数相同）
    daily_sales = daily_sales.loc[
        daily_sales["order-status"].isin(["Pending", "Shipped", "Unshipped"])
    ]
    daily_sales = (
        daily_sales.groupby(["sku"])[["quantity", "item-price"]].sum().reset_index()
    )
    daily_sales = daily_sales.rename(
        columns={"sku": "SKU", "quantity": "订单量", "item-price": "销售额"}
    )

    ad_columns_keep = [
        "广告SKU",
        "广告ASIN",
        "展示量",
        "点击量",
        "花费",
        "7天总销售量(#)",
    ]
    daily_ad_report = daily_ad_report[ad_columns_keep]
    daily_ad_report = daily_ad_report.rename(
        columns={
            "广告SKU": "SKU",
            "广告ASIN": "ASIN",
            "花费": "广告花费",
            "展示量": "曝光量",
            "7天总销售量(#)": "广告订单",
        }
    )
    daily_ad_report = (
        daily_ad_report.groupby(["SKU"])
        .agg(
            {
                "曝光量": "sum",
                "点击量": "sum",
                "广告花费": "sum",
                "广告订单": "sum",
                "ASIN": "first",
            }
        )
        .reset_index()
    )
    daily_ad_report["单次点击花费"] = (
        daily_ad_report["广告花费"] / daily_ad_report["点击量"]
        if daily_ad_report["点击量"].sum() > 0
        else 0
    )

    fba_columns_keep = ["sku", "available"]
    fba = fba[fba_columns_keep]
    fba = fba.rename(columns={"sku": "SKU", "available": "可售库存"})

    merged_data = pd.merge(daily_sales, daily_ad_report, on="SKU", how="outer")
    merged_data = pd.merge(merged_data, fba, on="SKU", how="outer")
    merged_data = merged_data.fillna(0)
    merged_data = merged_data.sort_values("SKU")

    overview_data = {
        "日期(US)": [],
        "SKU": [],
        "ASIN": [],
        "订单量": [],
        "销售额": [],
        "曝光量": [],
        "点击量": [],
        "单次点击花费": [],
        "广告花费": [],
        "广告订单": [],
        "可售库存": [],
    }
    df_overview = pd.DataFrame(overview_data)
    df_overview = pd.concat([df_overview, merged_data], ignore_index=True)
    df_overview = df_overview.fillna(0)
    df_overview["日期(US)"] = report_date

    for column in ["订单量", "曝光量", "点击量", "广告订单", "可售库存"]:
        df_overview[column] = df_overview[column].astype(int)
    for column in ["单次点击花费", "广告花费"]:
        df_overview[column] = df_overview[column].round(2)

    # 模板文件路径
    template_file = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "apps",
        "model_file",
        "daily_template.xlsx",
    )
    # 日报文件路径
    project_daily_file_path = os.path.join(
        project_folder_path, f"{project_name}_{report_date}_日报.xlsx"
    )
    # 加载模板文件
    workbook = load_workbook(template_file)
    ws = workbook.active

    for r_idx, row in enumerate(
        dataframe_to_rows(df_overview, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    workbook.save(project_daily_file_path)

    with open(project_daily_file_path, "rb") as f:
        file_content = f.read()

    return file_content, f"{project_name}_{report_date}_日报.xlsx"


def process_daily_report(
    project_name, report_date, sales_report, ad_report, fba_report
):
    current_time = datetime.datetime.now().strftime("%H-%M-%S")
    source_folder = os.getcwd()
    os.chdir(source_folder)

    project_folder_path = os.path.join(source_folder, "project", project_name, "日报")
    os.makedirs(project_folder_path, exist_ok=True)

    tmp_folder_path = os.path.join(source_folder, "project", project_name, "tmp")
    os.makedirs(tmp_folder_path, exist_ok=True)

    sales_report_path = f"{project_name}_sales_report_{report_date}.txt"
    ad_report_path = f"{project_name}_ad_report_{report_date}.xlsx"
    fba_report_path = f"{project_name}_fba_report_{report_date}.txt"

    with open(sales_report_path, "wb") as f:
        f.write(sales_report.read())

    with open(ad_report_path, "wb") as f:
        f.write(ad_report.read())

    with open(fba_report_path, "wb") as f:
        f.write(fba_report.read())

    daily_sales = pd.read_csv(sales_report_path, sep="\t", encoding="utf-8")
    daily_ad_report = pd.read_excel(ad_report_path, engine="openpyxl")
    fba = pd.read_csv(fba_report_path, sep="\t", encoding="utf-8")

    files_to_copy = [
        (
            sales_report_path,
            f"{tmp_folder_path}/{project_name}_sales_report_{report_date}_{current_time}.txt",
        ),
        (
            ad_report_path,
            f"{tmp_folder_path}/{project_name}_ad_report_{report_date}_{current_time}.xlsx",
        ),
        (
            fba_report_path,
            f"{tmp_folder_path}/{project_name}_fba_report_{report_date}_{current_time}.txt",
        ),
    ]

    for src, dst in files_to_copy:
        shutil.copy(src, dst)
        os.remove(src)

    daily_sales = daily_sales.loc[
        daily_sales["order-status"].isin(["Pending", "Shipped", "Unshipped"])
    ]
    daily_sales = (
        daily_sales.groupby(["sku"])[["quantity", "item-price"]].sum().reset_index()
    )
    daily_sales = daily_sales.rename(
        columns={"sku": "SKU", "quantity": "订单量", "item-price": "销售额"}
    )

    ad_columns_keep = [
        "广告SKU",
        "广告ASIN",
        "展示量",
        "点击量",
        "花费",
        "7天总销售量(#)",
    ]
    daily_ad_report = daily_ad_report[ad_columns_keep]
    daily_ad_report = daily_ad_report.rename(
        columns={
            "广告SKU": "SKU",
            "广告ASIN": "ASIN",
            "花费": "广告花费",
            "展示量": "曝光量",
            "7天总销售量(#)": "广告订单",
        }
    )
    daily_ad_report = (
        daily_ad_report.groupby(["SKU"])
        .agg(
            {
                "曝光量": "sum",
                "点击量": "sum",
                "广告花费": "sum",
                "广告订单": "sum",
                "ASIN": "first",
            }
        )
        .reset_index()
    )
    daily_ad_report["单次点击花费"] = (
        daily_ad_report["广告花费"] / daily_ad_report["点击量"]
        if daily_ad_report["点击量"].sum() > 0
        else 0
    )

    fba_columns_keep = ["sku", "available"]
    fba = fba[fba_columns_keep]
    fba = fba.rename(columns={"sku": "SKU", "available": "可售库存"})

    merged_data = pd.merge(daily_sales, daily_ad_report, on="SKU", how="outer")
    merged_data = pd.merge(merged_data, fba, on="SKU", how="outer")
    merged_data = merged_data.fillna(0)
    # 按SKU列排序
    merged_data = merged_data.sort_values("SKU")

    overview_data = {
        "日期(US)": [],
        "SKU": [],
        "ASIN": [],
        "订单量": [],
        "销售额": [],
        "曝光量": [],
        "点击量": [],
        "单次点击花费": [],
        "广告花费": [],
        "广告订单": [],
        "可售库存": [],
    }
    df_overview = pd.DataFrame(overview_data)
    df_overview = pd.concat([df_overview, merged_data], ignore_index=True)
    df_overview = df_overview.fillna(0)
    df_overview["日期(US)"] = report_date

    for column in ["订单量", "曝光量", "点击量", "广告订单", "可售库存"]:
        df_overview[column] = df_overview[column].astype(int)
    for column in ["单次点击花费", "广告花费"]:
        df_overview[column] = df_overview[column].round(2)
    # 模板文件路径
    template_file = os.path.join(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
        "apps",
        "model_file",
        "daily_template.xlsx",
    )
    # 日报文件路径
    project_daily_file_path = os.path.join(
        project_folder_path, f"{project_name}_{report_date}_日报.xlsx"
    )
    # 加载模板文件
    workbook = load_workbook(template_file)
    ws = workbook.active

    for r_idx, row in enumerate(
        dataframe_to_rows(df_overview, index=False, header=True), 1
    ):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(
                horizontal="center", vertical="center"
            )

    workbook.save(project_daily_file_path)

    with open(project_daily_file_path, "rb") as f:
        file_content = f.read()

    return file_content, f"{project_name}_{report_date}_日报.xlsx"


@daily_report_bp.route("/daily-report", methods=["GET", "POST"])
def daily_report():
    if request.method == "POST":
        project_name = request.form.get("project_name")
        report_date = request.form.get("report_date")

        # 优先使用已上传的文件路径
        sales_report_path = request.form.get("sales_report_path")
        ad_report_path = request.form.get("ad_report_path")
        fba_report_path = request.form.get("fba_report_path")

        # 如果没有已上传的文件路径，则使用传统方式
        sales_report_file = request.files.get("sales_report")
        ad_report_file = request.files.get("ad_report")
        fba_report_file = request.files.get("fba_report")

        # 验证必需参数
        if not project_name or not report_date:
            flash("请填写项目名称和报表日期")
            return redirect(url_for("dataset.daily_report_page"))

        # 检查是否有文件路径或文件上传
        has_paths = sales_report_path and ad_report_path and fba_report_path
        has_files = sales_report_file and ad_report_file and fba_report_file

        if not (has_paths or has_files):
            flash("请上传所有文件或确保所有文件已上传完成")
            return redirect(url_for("dataset.daily_report_page"))

        # 验证文件格式（如果有文件上传）
        if has_files:
            if not (
                allowed_file(sales_report_file.filename)
                and allowed_file(ad_report_file.filename)
                and allowed_file(fba_report_file.filename)
            ):
                flash("文件格式不正确")
                return redirect(url_for("dataset.daily_report_page"))

        try:
            # 如果有文件路径，使用路径处理；否则使用上传的文件
            if has_paths:
                # 验证文件是否存在
                if not all(
                    os.path.exists(path)
                    for path in [sales_report_path, ad_report_path, fba_report_path]
                ):
                    flash("文件不存在，请重新上传")
                    return redirect(url_for("dataset.daily_report_page"))

                file_content, filename = process_daily_report_from_paths(
                    project_name,
                    report_date,
                    sales_report_path,
                    ad_report_path,
                    fba_report_path,
                )
            else:
                file_content, filename = process_daily_report(
                    project_name,
                    report_date,
                    sales_report_file,
                    ad_report_file,
                    fba_report_file,
                )

            # 记录生成日报成功日志
            LogService.log(
                action="生成日报",
                resource="日报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 文件: {filename}",
                log_type="user",
                level="info",
            )

            # 手动创建响应以避免Flask send_file的中文文件名问题
            from flask import make_response
            import urllib.parse

            response = make_response(file_content)
            response.headers["Content-Type"] = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            # 只使用filename*参数，不使用filename参数，避免浏览器选择截断的文件名
            encoded_filename = urllib.parse.quote(filename)
            response.headers["Content-Disposition"] = (
                f"attachment; filename*=UTF-8''{encoded_filename}"
            )

            return response
        except Exception as e:
            # 记录生成日报失败日志
            LogService.log(
                action="生成日报失败",
                resource="日报功能",
                details=f"项目: {project_name}, 日期: {report_date}, 错误: {str(e)}",
                log_type="user",
                level="error",
            )
            flash(f"生成日报时发生错误: {str(e)}", "error")
            return redirect(url_for("dataset.daily_report_page"))

    return render_template("data-analysis/daily_report.html")
